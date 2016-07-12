from openpyxl import Workbook
from openpyxl.styles import Font, colors
import MySQLdb as mysql
from settings import DB_HOST, DB_NAME, DB_USER, DB_PASSWD

conn = mysql.connect(host=DB_HOST, user=DB_USER, passwd=DB_PASSWD, db=DB_NAME, use_unicode=True, charset='utf8')

cursor = conn.cursor()


sql_code = 'SELECT leagues.id, leagues.name, leagues.bookmaker, leagues.sport, sports.name, ' \
           'sportsmatch.uuid as sportuuid, leaguesmatch.uuid as uuid, bookmakers.name as bk ' \
           'FROM leagues ' \
           'LEFT JOIN leaguesmatch ON leagues.id = leaguesmatch.league ' \
           'LEFT JOIN sportsmatch ON leagues.sport = sportsmatch.sport ' \
           'LEFT JOIN sports ON leagues.sport = sports.id ' \
           'LEFT JOIN bookmakers ON leagues.bookmaker = bookmakers.id ' \
           'ORDER BY sportuuid '

sql_code2 = 'SELECT * FROM leaguesmatch ' \
           'ORDER BY uuid'

cursor.execute(sql_code)
rows = cursor.fetchall()

cursor.execute(sql_code2)
matches = cursor.fetchall()

wb = Workbook()
sport_sheets = {}
bookmakers = {}

# Заполнение лиг которые есть в базе

ws = wb.active
for row in rows:

    # Определение страницы
    uuid_ = row[5]
    if uuid_ in sport_sheets:
        ws = wb.get_sheet_by_name(sport_sheets[uuid_])
    else:
        sportname = row[4]
        ws = wb.create_sheet(sportname)
        sport_sheets[uuid_] = sportname

    # Определение столбика
    bk_id = row[2]
    bk_name = row[7]

    j = 2
    while ws.cell(row=1, column=j).value is not None:
        if ws.cell(row=1, column=j).value == bk_id:
            break
        j += 2

    bk_id_cell = ws.cell(row=1, column=j)
    if bk_id_cell.value is None:
        bk_id_cell.value = bk_id
        bk_name_cell = ws.cell(row=1, column=j-1)
        bk_name_cell.value = bk_name

    # Определение строки
    i = 2
    while ws.cell(row=i, column=j).value is not None:
        i += 1

    id_cell = ws.cell(row=i, column=j)
    name_cell = ws.cell(row=i, column=j-1)

    not_matched = row[6] is None

    id_cell.value = row[0]
    name_cell.value = row[1]

    if not_matched:
        id_cell.font = Font(color=colors.RED)


