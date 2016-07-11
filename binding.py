import MySQLdb as mysql
from distance import distance
from settings import DB_HOST, DB_NAME, DB_USER, DB_PASSWD, LIVENSHTAIN_MIN
from uuid import uuid4
from openpyxl import Workbook
from openpyxl.styles import Font, colors


class Matched():

    elements = []

    def get(self, match):
        """
        Проверяет вхождение
        :rtype: bool
        :type match: tuple
        """
        for el in self.elements:
            if el[0] == match[0] and el[1] == match[1] or \
                                    el[0] == match[1] and el[1] == match[0]:
                return el
        return None


def matching(matched_table, matching_table, match_field):

    conn = mysql.connect(host=DB_HOST, user=DB_USER, passwd=DB_PASSWD, db=DB_NAME, use_unicode=True, charset='utf8')

    cursor = conn.cursor()

    sql_code = 'SELECT {0}.id as id, {0}.name as name, {0}.bookmaker as bookmaker FROM {0} ' \
               'LEFT JOIN {1} ' \
               'ON {0}.id = {1}.{2} ' \
               'WHERE {1}.uuid IS NULL'.format(matched_table, matching_table, match_field)

    cursor.execute(sql_code)

    rows_without_matching = cursor.fetchall()

    print("Для {0} строк не найдены соответствия в таблице соответствий".format(str(len(rows_without_matching))))

    matched = Matched()
    for sport1 in rows_without_matching:
        for sport2 in rows_without_matching:
            if sport1[2] != sport2[2]:  # Если строчки для разных букмекеров
                dist = distance(sport1[1], sport2[1])
                match = (sport1[0], sport2[0], dist)
                existed_match = matched.get(match)
                if existed_match is None:
                    matched.elements.append(match)
                else:
                    if existed_match[2] > match[2]:
                        matched.elements.remove(existed_match)
                        matched.elements.append(match)

    print("Установлено {0} соответствий с тем или иным расстоянием левинштейна".format(
        str(
            len(matched.elements)
        )
    ))

    matched_list = []
    for el in matched.elements:
        if el[2] <= LIVENSHTAIN_MIN:
            matched_list.append((el[0], el[1]))

    print("Только {0} соответствий обладают расстоянием левенштейна менее {1}".format(
        str(len(matched_list)),
        str(LIVENSHTAIN_MIN)
    ))

    list_for_insert = []
    for el in matched_list:
        uuid = uuid4().bytes
        list_for_insert.append((el[0], uuid))
        list_for_insert.append((el[1], uuid))

    sql_code = 'INSERT INTO {0} (`{1}`, `uuid`) VALUES (%s, %s)'.format(matching_table, match_field)

    cursor.executemany(sql_code, list_for_insert)

    conn.commit()

    conn.close()


# print("Поехали, спорты!")
# matching("sports", "sportsmatch", "sport")
#
# print("Поехали, лиги!!!111!")
# matching("leagues", "leaguesmatch", "league")
#
# print("Поехали, УЧАСТНИКИ!!!11111!")
# matching("participants", "participantsmatch", "participant")


conn = mysql.connect(host=DB_HOST, user=DB_USER, passwd=DB_PASSWD, db=DB_NAME, use_unicode=True, charset='utf8')

cursor = conn.cursor()

sql_code = 'SELECT leagues.id, leagues.name, leagues.bookmaker, leagues.sport, sports.name, ' \
           'sportsmatch.uuid as sportuuid, leaguesmatch.uuid as uuid, bookmakers.name as bk ' \
           'FROM leagues ' \
           'LEFT JOIN leaguesmatch ON leagues.id = leaguesmatch.league ' \
           'LEFT JOIN sportsmatch ON leagues.sport = sportsmatch.sport ' \
           'LEFT JOIN sports ON leagues.sport = sports.id ' \
           'LEFT JOIN bookmakers ON leagues.bookmaker = bookmakers.id ' \
           'ORDER BY sportuuid '

cursor.execute(sql_code)
rows = cursor.fetchall()

wb = Workbook()
sport_sheets = {}
bookmakers = {}

ws = wb.active
for row in rows:

    # Определение страницы
    uuid_ = row[5]
    if uuid_ in sport_sheets:
        ws = wb.get_sheet_by_name(sport_sheets[uuid_])
    else:
        sportname = row[4]
        ws = wb.create_sheet(sportname)
        sport_sheets[uuid_] = sportname

    # Определение столбика
    bk_id = row[2]
    bk_name = row[7]

    j = 2
    while ws.cell(row=1, column=j).value is not None:
        if ws.cell(row=1, column=j).value == bk_id:
            break
        j += 2

    bk_id_cell = ws.cell(row=1, column=j)
    if bk_id_cell.value is None:
        bk_id_cell.value = bk_id
        bk_name_cell = ws.cell(row=1, column=j-1)
        bk_name_cell.value = bk_name

    # Определение строки
    i = 2
    while ws.cell(row=i, column=j).value is not None:
        i += 1

    id_cell = ws.cell(row=i, column=j)
    name_cell = ws.cell(row=i, column=j-1)

    not_matched = row[6] is None

    id_cell.value = row[0]
    name_cell.value = row[1]

    if not_matched:
        id_cell.font = Font(color=colors.RED)

wb.save("simple.xslx")

























wb.sheetnames
wb.save("sample.xlsx")